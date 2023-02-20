from tempfile import NamedTemporaryFile

import pandas as pd
from pydrive2.auth import GoogleAuth
from pydrive2.drive import GoogleDrive
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter


class ExcelWorker:
    def __init__(self):
        self._auth()

    def _auth(self):
        self._gauth = GoogleAuth()
        self._gauth.LoadCredentialsFile("credentials.json")
        if self._gauth.credentials is None:
            # Authenticate if they're not there
            self._gauth.LocalWebserverAuth()
        elif self._gauth.access_token_expired:
            # Refresh them if expired
            self._gauth.Refresh()
        else:
            # Initialize the saved creds
            self._gauth.Authorize()
        self._gauth.SaveCredentialsFile("credentials.json")
        self._drive = GoogleDrive(self._gauth)

    def savedf(self, df: pd.DataFrame, path_file: str):
        with NamedTemporaryFile() as tmp:
            wb = Workbook()

            writer = pd.ExcelWriter(tmp.name, engine="openpyxl")
            writer.book = wb
            sheet_name = "Sheet"
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            for col_ind in range(1, df.shape[1] + 1):
                writer.sheets[sheet_name].column_dimensions[
                    get_column_letter(col_ind)
                ].width = 40.0  # вынести отдельно передача стилей

            wb.save(tmp.name)

            f = self._drive.CreateFile({"id": path_file})
            f.SetContentFile(tmp.name)
            f["title"] = "test.xlsx"  # вынести
            f.Upload()
